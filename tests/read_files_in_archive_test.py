import os
import zipfile
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from scripts.create_archive import create_archive
import pytest


path_to_files = os.path.join(os.path.dirname(os.path.abspath(__file__)), os.path.pardir, 'files')
path_resources = os.path.join(os.path.dirname(os.path.abspath(__file__)),os.path.pardir, 'resources')
path_zip = os.path.join(path_resources, "QAguru3_6.zip")


def test_contains_csv_in_archive():
    create_archive(path_to_files, path_zip)
    with zipfile.ZipFile(path_zip) as zf:
        cf = zf.extract("csv_file.csv")

        with open(cf) as csvfile:
            list_csv = []
            csvfile = csv.reader(csvfile)
            for r in csvfile:
                text_in_csv = '; '.join(r)
                list_csv.append(text_in_csv)
            assert list_csv[1] == 'booker12;9012;Rachel;Booker', f'Текст в в файле .csv не совпадает\n' \
                                                                 f'Ожидаемый результат: {"booker12;9012;Rachel;Booker"}\n' \
                                                                 f'Фактический результат: {list_csv[1]}\n'
    os.remove(cf)


def test_contains_pdf_in_archive():
    create_archive(path_to_files, path_zip)
    with zipfile.ZipFile(path_zip) as zf:
        cf = zf.extract("pdf_file.pdf")

        reader = PdfReader(cf)
        page = reader.pages[0]
        text = page.extract_text()
        assert 'Здесь будет' in text, f'Текст: {"Здесь будет"} не содержится в {text}'
        os.remove(cf)


def test_contains_xlsx_in_archive():
    create_archive(path_to_files, path_zip)
    with zipfile.ZipFile(path_zip) as zf:
        cf = zf.extract("excel_file.xlsx")

        workbook = load_workbook(cf)
        sheet = workbook.active
        name_user = sheet.cell(row=4, column=2).value
        assert name_user == 'Philip', 'Значение колонки в таблице не совпадает!'
        sheet = workbook.close()
    os.remove(cf)
